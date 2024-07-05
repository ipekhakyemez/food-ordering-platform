import openpyxl

def read_menu_data(sheet_name):
    try:
        workbook = openpyxl.load_workbook("data/menu.xlsx")
        sheet = workbook[sheet_name]
        rows = sheet.max_row
        
        menu_data = []
        for row in range(2, rows + 1):
            product_name = sheet.cell(row, 1).value
            product_price = sheet.cell(row, 2).value
            menu_data.append((product_name, product_price))
        
        return menu_data
    
    except Exception as e:
        print(f"Hata oluştu: {str(e)}")
        return None
